import hashlib
import hmac
import io
import os
import re
import secrets
import tempfile
import unicodedata
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import httpx
from sqlmodel import Session, select
from sqlalchemy import text as sa_text

from .database import create_db, get_session, engine
from .facturapi import get_client
from .models import ComplementoPago, Factura, Partida, Pedimento, Producto, User
from .parser import parse_pedimento

load_dotenv(Path(__file__).resolve().parent.parent / ".env")

app = FastAPI(title="Analizador de Pedimentos")

limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# ── Auth helpers ──────────────────────────────────────────────────────────────

_TOKEN_TTL = timedelta(hours=24)
_active_tokens: dict[str, tuple[str, datetime]] = {}  # token -> (email, expires_at)
_bearer = HTTPBearer(auto_error=False)

_INITIAL_USER_EMAIL = os.environ.get("INITIAL_USER_EMAIL", "").strip().lower()
_INITIAL_USER_PASSWORD = os.environ.get("INITIAL_USER_PASSWORD", "")


def _hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 260_000)
    return f"{salt}${dk.hex()}"


def _verify_password(password: str, hashed: str) -> bool:
    try:
        salt, dk_hex = hashed.split("$", 1)
    except ValueError:
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 260_000)
    return hmac.compare_digest(dk.hex(), dk_hex)


def _seed_users(session: Session) -> None:
    if not _INITIAL_USER_EMAIL or not _INITIAL_USER_PASSWORD:
        return
    existing = session.exec(select(User).where(User.email == _INITIAL_USER_EMAIL)).first()
    if not existing:
        session.add(User(email=_INITIAL_USER_EMAIL, hashed_password=_hash_password(_INITIAL_USER_PASSWORD)))
        session.commit()


def _prune_tokens() -> None:
    now = datetime.now(timezone.utc)
    expired = [t for t, (_, exp) in _active_tokens.items() if exp <= now]
    for t in expired:
        del _active_tokens[t]


_PUBLIC_PATHS = {"/auth/login", "/"}

def require_auth(
    credentials: HTTPAuthorizationCredentials | None = Depends(_bearer),
) -> str:
    token = credentials.credentials if credentials else None
    if not token or token not in _active_tokens:
        raise HTTPException(status_code=401, detail="No autenticado")
    email, expires_at = _active_tokens[token]
    if datetime.now(timezone.utc) >= expires_at:
        del _active_tokens[token]
        raise HTTPException(status_code=401, detail="Sesión expirada")
    return email


@app.middleware("http")
async def auth_middleware(request: Request, call_next):
    path = request.url.path
    if path in _PUBLIC_PATHS or path.startswith("/static/") or path.startswith("/assets/"):
        return await call_next(request)
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.removeprefix("Bearer ").strip() if auth_header.startswith("Bearer ") else None
    if not token or token not in _active_tokens:
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=401, content={"detail": "No autenticado"})
    _, expires_at = _active_tokens[token]
    if datetime.now(timezone.utc) >= expires_at:
        del _active_tokens[token]
        from fastapi.responses import JSONResponse
        return JSONResponse(status_code=401, content={"detail": "Sesión expirada"})
    return await call_next(request)

FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"
MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB


# ── SAT search helpers ────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    s = s.lower()
    s = unicodedata.normalize("NFD", s)
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _build_fts_query(words: list, mode: str = "AND") -> str:
    tokens = [f'"{w}"*' for w in words]
    return (" OR " if mode == "OR" else " ").join(tokens)


def _fts_search(session: Session, fts_table: str, base_table: str,
                words: list, limit: int, prefer_or: bool = False) -> list:
    """FTS5 search. prefer_or=True uses OR directly (good for live autocomplete).
    prefer_or=False tries AND first, falls back to OR (better precision for batch)."""
    def _run(mode: str):
        return session.execute(sa_text(f"""
            SELECT c.key, c.description
            FROM {fts_table} f JOIN {base_table} c ON c.rowid = f.rowid
            WHERE {fts_table} MATCH :q
            ORDER BY bm25({fts_table}, 0, 1) LIMIT :lim
        """), {"q": _build_fts_query(words, mode), "lim": limit}).fetchall()

    try:
        if prefer_or:
            rows = _run("OR")
        else:
            rows = _run("AND")
            if not rows and len(words) > 1:
                rows = _run("OR")
    except Exception:
        longest = max(words, key=len)
        rows = session.execute(sa_text(
            f"SELECT key, description FROM {base_table} "
            "WHERE description LIKE :q LIMIT :lim"
        ), {"q": f"%{longest}%", "lim": limit}).fetchall()

    return [{"key": r[0], "description": r[1]} for r in rows]


# ── HS chapter context ────────────────────────────────────────────────────────

HS_CHAPTERS = {
    # Sección I — Animales vivos y productos del reino animal
    "01": "Animales vivos",
    "02": "Carne y despojos comestibles",
    "03": "Pescado, crustáceos y moluscos",
    "04": "Lácteos, huevos, miel y productos comestibles de origen animal",
    "05": "Los demás productos de origen animal",
    # Sección II — Productos del reino vegetal
    "06": "Plantas vivas y productos de la floricultura",
    "07": "Hortalizas, plantas, raíces y tubérculos",
    "08": "Frutas y frutos comestibles",
    "09": "Café, té, yerba mate y especias",
    "10": "Cereales",
    "11": "Productos de la molinería, malta, almidón",
    "12": "Semillas oleaginosas, plantas industriales",
    "13": "Gomas, resinas y demás jugos y extractos vegetales",
    "14": "Materias trenzables y demás productos de origen vegetal",
    # Sección III — Grasas y aceites
    "15": "Grasas y aceites animales, vegetales o microbianos",
    # Sección IV — Alimentos, bebidas, tabaco
    "16": "Preparaciones de carne, pescado o crustáceos",
    "17": "Azúcares y artículos de confitería",
    "18": "Cacao y sus preparaciones",
    "19": "Preparaciones a base de cereales, harina o almidón",
    "20": "Preparaciones de hortalizas, frutas y plantas",
    "21": "Preparaciones alimenticias diversas",
    "22": "Bebidas, líquidos alcohólicos y vinagre",
    "23": "Residuos y desperdicios de industrias alimentarias",
    "24": "Tabaco y sucedáneos del tabaco",
    # Sección V — Productos minerales
    "25": "Sal, azufre, tierras, piedras, yesos, cales y cementos",
    "26": "Minerales metalíferos, escorias y cenizas",
    "27": "Combustibles y aceites minerales (petróleo, gas, carbón)",
    # Sección VI — Productos de las industrias químicas
    "28": "Productos químicos inorgánicos",
    "29": "Productos químicos orgánicos",
    "30": "Productos farmacéuticos",
    "31": "Abonos y fertilizantes",
    "32": "Extractos curtientes, colorantes, pinturas y barnices",
    "33": "Aceites esenciales, perfumería y cosméticos",
    "34": "Jabones, agentes de superficie, lubricantes, ceras y velas",
    "35": "Albuminoides, colas, enzimas",
    "36": "Pólvoras, explosivos, artículos de pirotecnia",
    "37": "Productos fotográficos y cinematográficos",
    "38": "Productos diversos de las industrias químicas",
    # Sección VII — Plásticos y caucho
    "39": "Plásticos y sus manufacturas",
    "40": "Caucho y sus manufacturas",
    # Sección VIII — Pieles, cueros y peletería
    "41": "Pieles (excepto la peletería) y cueros",
    "42": "Manufacturas de cuero, marroquinería, artículos de viaje",
    "43": "Peletería y confecciones de peletería",
    # Sección IX — Madera y corcho
    "44": "Madera, carbón vegetal y manufacturas de madera",
    "45": "Corcho y sus manufacturas",
    "46": "Manufacturas de espartería y cestería",
    # Sección X — Pasta de madera, papel y cartón
    "47": "Pasta de madera y material fibroso celulósico",
    "48": "Papel, cartón y sus manufacturas",
    "49": "Libros, impresos, prensa y manuscritos",
    # Sección XI — Materias textiles y sus manufacturas
    "50": "Seda",
    "51": "Lana y pelo fino u ordinario",
    "52": "Algodón",
    "53": "Demás fibras textiles vegetales, hilados de papel",
    "54": "Filamentos sintéticos o artificiales",
    "55": "Fibras sintéticas o artificiales discontinuas",
    "56": "Guata, fieltro, telas sin tejer, cuerdas y cordeles",
    "57": "Alfombras y demás revestimientos para el suelo",
    "58": "Tejidos especiales, encajes, tapicería, pasamanería",
    "59": "Telas impregnadas, recubiertas o estratificadas",
    "60": "Tejidos de punto",
    "61": "Prendas y complementos de vestir de punto",
    "62": "Prendas y complementos de vestir excepto de punto",
    "63": "Los demás artículos textiles confeccionados",
    # Sección XII — Calzado, sombrerería
    "64": "Calzado, polainas y artículos análogos",
    "65": "Sombreros, tocados y sus partes",
    "66": "Paraguas, sombrillas, bastones y artículos similares",
    "67": "Plumas preparadas, flores artificiales, manufacturas de cabello",
    # Sección XIII — Piedra, cerámica, vidrio
    "68": "Manufacturas de piedra, yeso, cemento, amianto y mica",
    "69": "Productos cerámicos",
    "70": "Vidrio y sus manufacturas",
    # Sección XIV — Perlas, piedras preciosas y metales preciosos
    "71": "Perlas, piedras preciosas, metales preciosos y bisutería",
    # Sección XV — Metales comunes y sus manufacturas
    "72": "Fundición, hierro y acero",
    "73": "Manufacturas de fundición, hierro o acero",
    "74": "Cobre y sus manufacturas",
    "75": "Níquel y sus manufacturas",
    "76": "Aluminio y sus manufacturas",
    "78": "Plomo y sus manufacturas",
    "79": "Cinc y sus manufacturas",
    "80": "Estaño y sus manufacturas",
    "81": "Los demás metales comunes y sus manufacturas",
    "82": "Herramientas, artículos de cuchillería de metal común",
    "83": "Manufacturas diversas de metal común",
    # Sección XVI — Máquinas y aparatos
    "84": "Reactores nucleares, calderas, máquinas y aparatos mecánicos",
    "85": "Máquinas, aparatos y material eléctrico y electrónico",
    # Sección XVII — Material de transporte
    "86": "Vehículos y material para vías férreas",
    "87": "Vehículos automóviles, tractores, ciclos y demás terrestres",
    "88": "Aeronaves, vehículos espaciales y sus partes",
    "89": "Barcos y embarcaciones",
    # Sección XVIII — Instrumentos de precisión y médicos
    "90": "Instrumentos de óptica, fotografía, medida, medicina y cirugía",
    "91": "Aparatos de relojería",
    "92": "Instrumentos musicales",
    # Sección XIX — Armas y municiones
    "93": "Armas, municiones y sus partes",
    # Sección XX — Manufacturas diversas
    "94": "Muebles, mobiliario médico-quirúrgico, artículos de cama",
    "95": "Juguetes, juegos y artículos de deporte",
    "96": "Manufacturas diversas (bolígrafos, peines, botones, encendedores, pipas)",
    # Sección XXI — Objetos de arte
    "97": "Objetos de arte, de colección o de antigüedad",
}


def _chapter_hint(fraccion: str) -> str:
    ch = fraccion[:2]
    desc = HS_CHAPTERS.get(ch, "")
    return f"[Cap.{ch}: {desc}]" if desc else f"[Cap.{ch}]"


@app.on_event("startup")
def on_startup():
    create_db()
    with Session(engine) as session:
        _seed_users(session)


# ── Auth endpoints ────────────────────────────────────────────────────────────

@app.post("/auth/login")
@limiter.limit("5/minute")
def login(request: Request, body: dict, session: Session = Depends(get_session)):
    email = (body.get("email") or "").strip().lower()
    password = body.get("password") or ""
    user = session.exec(select(User).where(User.email == email)).first()
    if not user or not _verify_password(password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    token = secrets.token_hex(32)
    expires_at = datetime.now(timezone.utc) + _TOKEN_TTL
    _active_tokens[token] = (email, expires_at)
    _prune_tokens()
    return {"token": token, "email": email}


@app.post("/auth/logout")
def logout(_: str = Depends(require_auth), credentials: HTTPAuthorizationCredentials | None = Depends(_bearer)):
    token = credentials.credentials if credentials else None
    if token and token in _active_tokens:
        del _active_tokens[token]
    return {"ok": True}


@app.get("/auth/me")
def me(email: str = Depends(require_auth)):
    return {"email": email}


# ── Pedimentos ────────────────────────────────────────────────────────────────

@app.post("/parse")
async def parse(
    file: UploadFile = File(...),
    _: str = Depends(require_auth),
    session: Session = Depends(get_session),
):
    content = await file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "El archivo excede el tamaño máximo permitido (20 MB)")
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Solo se aceptan archivos PDF")

    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        result = parse_pedimento(tmp_path)
    except Exception as e:
        raise HTTPException(422, f"Error al procesar el PDF: {e}")
    finally:
        Path(tmp_path).unlink(missing_ok=True)

    # Deduplication: return existing record if same pedimento_num already stored
    existing = session.exec(
        select(Pedimento).where(Pedimento.pedimento_num == result["pedimento_num"])
    ).first()
    if existing:
        result["id"] = existing.id
        result["_duplicate"] = True
        result["dta"] = existing.dta if existing.dta is not None else result.get("dta")
        result["igi"] = existing.igi if existing.igi is not None else result.get("igi")
        result["prv"] = existing.prv if existing.prv is not None else result.get("prv")
        result["partidas"] = [
            {
                "id": p.id, "sec": p.sec, "fraccion": p.fraccion,
                "descripcion": p.descripcion, "cantidad": p.cantidad,
                "val_aduana": p.val_aduana, "val_comercial": p.val_comercial,
                "precio_unitario": p.precio_unitario,
                "tiene_incrementables": p.tiene_incrementables,
            }
            for p in existing.partidas
        ]
        return result

    # Persist to DB
    pedimento = Pedimento(
        pedimento_num=result["pedimento_num"],
        importador=result["importador"],
        tipo_cambio=result["tipo_cambio"],
        pdf_filename=file.filename,
        dta=result.get("dta"),
        igi=result.get("igi"),
        prv=result.get("prv"),
    )
    session.add(pedimento)
    session.flush()  # get pedimento.id before adding partidas

    for p in result["partidas"]:
        session.add(Partida(
            pedimento_id=pedimento.id,
            sec=p["sec"],
            fraccion=p["fraccion"],
            descripcion=p["descripcion"],
            cantidad=p["cantidad"],
            val_aduana=p["val_aduana"],
            val_comercial=p["val_comercial"],
            precio_unitario=p["precio_unitario"],
            tiene_incrementables=p["tiene_incrementables"],
        ))

    session.commit()
    result["id"] = pedimento.id
    return result


@app.get("/pedimentos")
def list_pedimentos(_: str = Depends(require_auth), session: Session = Depends(get_session)):
    from sqlalchemy import text as sa_text, func
    rows = session.exec(select(Pedimento).order_by(Pedimento.fecha_upload.desc())).all()
    counts = dict(session.execute(
        sa_text("SELECT pedimento_id, COUNT(*) FROM partida GROUP BY pedimento_id")
    ).fetchall())
    return [
        {
            "id": p.id,
            "pedimento_num": p.pedimento_num,
            "importador": p.importador,
            "tipo_cambio": p.tipo_cambio,
            "pdf_filename": p.pdf_filename,
            "fecha_upload": p.fecha_upload,
            "num_partidas": counts.get(p.id, 0),
        }
        for p in rows
    ]


@app.get("/pedimentos/{pedimento_id}")
def get_pedimento(pedimento_id: int, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    pedimento = session.get(Pedimento, pedimento_id)
    if not pedimento:
        raise HTTPException(404, "Pedimento no encontrado")
    return {
        "id": pedimento.id,
        "pedimento_num": pedimento.pedimento_num,
        "importador": pedimento.importador,
        "tipo_cambio": pedimento.tipo_cambio,
        "pdf_filename": pedimento.pdf_filename,
        "fecha_upload": pedimento.fecha_upload,
        "dta": pedimento.dta,
        "igi": pedimento.igi,
        "prv": pedimento.prv,
        "partidas": [
            {
                "id": p.id,
                "sec": p.sec,
                "fraccion": p.fraccion,
                "descripcion": p.descripcion,
                "cantidad": p.cantidad,
                "val_aduana": p.val_aduana,
                "val_comercial": p.val_comercial,
                "precio_unitario": p.precio_unitario,
                "tiene_incrementables": p.tiene_incrementables,
            }
            for p in pedimento.partidas
        ],
    }


@app.post("/pedimentos/{pedimento_id}/automap")
async def automap_claves(pedimento_id: int, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    from google import genai
    from google.genai import types
    import json

    pedimento = session.get(Pedimento, pedimento_id)
    if not pedimento:
        raise HTTPException(404, "Pedimento no encontrado")

    # Deduplicate by fraccion — one mapping per fraccion is enough
    seen = {}
    for p in pedimento.partidas:
        if p.fraccion not in seen:
            seen[p.fraccion] = p
    partidas = list(seen.values())

    # Skip fracciones already mapped
    already = {p.fraccion for p in session.exec(select(Producto)).all()}
    to_map = [p for p in partidas if p.fraccion not in already]
    if not to_map:
        return {"mapped": 0, "results": [], "message": "Todas las fracciones ya están mapeadas"}

    gemini = genai.Client(api_key=os.environ["GEMINI_API_KEY"])

    # ── Tool functions ────────────────────────────────────────────────────────
    def search_sat_catalog(query: str) -> list:
        words = [w for w in re.findall(r"[a-z]{3,}", _norm(query))]
        if not words:
            return []
        return _fts_search(session, "sat_claves_fts", "sat_claves", words, limit=25)

    def search_sat_units(query: str) -> list:
        words = [w for w in re.findall(r"[a-z]{2,}", _norm(query))]
        if not words:
            return []
        return _fts_search(session, "sat_unidades_fts", "sat_unidades", words, limit=15)

    # ── Tool declarations ─────────────────────────────────────────────────────
    combined_tool = types.Tool(function_declarations=[
        types.FunctionDeclaration(
            name="search_sat_catalog",
            description=(
                "Busca c_ClaveProdServ en el catálogo oficial SAT. "
                "Devuelve hasta 25 resultados ordenados por relevancia."
            ),
            parameters=types.Schema(
                type=types.Type.OBJECT,
                properties={
                    "query": types.Schema(
                        type=types.Type.STRING,
                        description=(
                            "Término de búsqueda en español formal, "
                            "ej: 'pitillo', 'funda aislante vaso', 'contenedor polipropileno'"
                        ),
                    )
                },
                required=["query"],
            ),
        ),
        types.FunctionDeclaration(
            name="search_sat_units",
            description=(
                "Busca c_ClaveUnidad SAT (unidad de medida para CFDI). "
                "Ej: 'pieza'→H87, 'kilogramo'→KGM, 'litro'→LTR, 'metro'→MTR."
            ),
            parameters=types.Schema(
                type=types.Type.OBJECT,
                properties={
                    "query": types.Schema(
                        type=types.Type.STRING,
                        description="Unidad de medida en español, ej: 'pieza', 'kilogramo', 'litro'",
                    )
                },
                required=["query"],
            ),
        ),
    ])

    # ── Shared config (with thinking budget) ─────────────────────────────────
    config = types.GenerateContentConfig(
        tools=[combined_tool],
        temperature=0,
        thinking_config=types.ThinkingConfig(thinking_budget=8192),
    )

    # ── Reusable agentic loop ─────────────────────────────────────────────────
    def run_loop(messages: list, system: str, n_items: int) -> list | None:
        cfg = types.GenerateContentConfig(
            system_instruction=system,
            tools=[combined_tool],
            temperature=0,
            thinking_config=types.ThinkingConfig(thinking_budget=8192),
        )
        parse_attempts = 0
        for _ in range(35):
            response = gemini.models.generate_content(
                model="gemini-2.5-flash",
                contents=messages,
                config=cfg,
            )
            candidate = response.candidates[0]
            messages.append(candidate.content)

            tool_calls = [p for p in candidate.content.parts if p.function_call]
            if tool_calls:
                tool_results = []
                for part in tool_calls:
                    fc = part.function_call
                    query = fc.args.get("query", "")
                    if fc.name == "search_sat_catalog":
                        results_data = search_sat_catalog(query)
                    elif fc.name == "search_sat_units":
                        results_data = search_sat_units(query)
                    else:
                        results_data = []
                    tool_results.append(types.Part.from_function_response(
                        name=fc.name, response={"results": results_data},
                    ))
                messages.append(types.Content(role="user", parts=tool_results))
                continue

            text_parts = [p.text for p in candidate.content.parts if hasattr(p, "text") and p.text]
            full_text = "\n".join(text_parts).strip()
            clean = re.sub(r'^```(?:json)?\s*|\s*```$', '', full_text, flags=re.MULTILINE).strip()
            match = re.search(r'\[.*\]', clean, re.DOTALL)
            if match:
                try:
                    return json.loads(match.group())
                except json.JSONDecodeError:
                    pass

            parse_attempts += 1
            if parse_attempts <= 2:
                messages.append(types.Content(role="user", parts=[types.Part(text=(
                    "Tu respuesta no contiene JSON válido. Necesito exactamente un array JSON con "
                    f"{n_items} objetos, claves: \"fraccion\", \"key\", \"unit_key\", \"description\", \"confidence\". "
                    "Sin texto adicional ni bloques de código. Inténtalo de nuevo."
                ))]))
                continue
            break
        return None

    # ── First pass — strict, minimum 3 searches per product ──────────────────
    def _items_text(partidas):
        return "\n".join(
            f'- fraccion={p.fraccion} {_chapter_hint(p.fraccion)} | "{p.descripcion}"'
            for p in partidas
        )

    system_pass1 = (
        "Eres un experto en clasificación SAT para CFDI 4.0 en México. "
        "Tienes dos herramientas: search_sat_catalog (c_ClaveProdServ) y search_sat_units (c_ClaveUnidad).\n"
        "REGLAS OBLIGATORIAS:\n"
        "(1) SIEMPRE usa search_sat_catalog — nunca inventes un código.\n"
        "(2) Para CADA producto busca MÍNIMO 3 VECES con términos distintos antes de considerar null: "
        "primero el término específico, luego un sinónimo, luego la categoría genérica del capítulo HS.\n"
        "(3) Para CADA producto usa search_sat_units para determinar la unidad correcta "
        "(pieza=H87, kilogramo=KGM, litro=LTR, metro=MTR, par=PR, caja=XBX, etc.).\n"
        "(4) El catálogo usa español formal: 'popote'→'pitillo'; 'plástico'→'polietileno','polipropileno'; "
        "'manga vaso'→'funda','aislante','protector'; 'portavaso'→'soporte','bandeja','porta'; "
        "'tapa domo'→'tapa','cubierta','tapadera'; 'contenedor'→'recipiente','envase'.\n"
        "(5) El capítulo HS entre corchetes indica la categoría — úsalo para refinar búsquedas.\n"
        "(6) null SOLO si después de 3+ búsquedas no encuentras absolutamente nada relacionado.\n"
        "(7) Para cada resultado incluye un campo confidence: "
        "'high' si el código es específico y claramente correcto para el producto; "
        "'medium' si es razonablemente cercano pero no exacto; "
        "'low' si es el más cercano disponible pero puede no ser correcto.\n"
        "(8) Solo responde JSON cuando hayas procesado TODOS los productos."
    )

    user_msg1 = (
        f"Clasifica estos {len(to_map)} productos con c_ClaveProdServ y c_ClaveUnidad SAT para CFDI.\n"
        f"La fracción arancelaria NO es el código SAT; el capítulo HS es solo contexto de categoría.\n\n"
        f"Productos:\n{_items_text(to_map)}\n\n"
        f"IMPORTANTE: busca cada producto AL MENOS 3 VECES con términos diferentes antes de poner null. "
        f"Responde ÚNICAMENTE con este JSON (sin markdown):\n"
        f'[{{"fraccion":"...","key":"... o null","unit_key":"H87 u otra","description":"... o null","confidence":"high|medium|low"}}]'
    )

    final_json = run_loop(
        [types.Content(role="user", parts=[types.Part(text=user_msg1)])],
        system_pass1,
        len(to_map),
    )
    if not final_json:
        raise HTTPException(500, "Gemini no devolvió un JSON válido con los códigos")

    # ── Second pass — rescue nulls with relaxed "best match" mandate ──────────
    null_fracciones = {
        item["fraccion"] for item in final_json
        if not item.get("key") or item["key"].lower() == "null"
    }
    if null_fracciones:
        null_partidas = [p for p in to_map if p.fraccion in null_fracciones]

        system_pass2 = (
            "Eres un experto en clasificación SAT para CFDI 4.0 en México. "
            "Tienes dos herramientas: search_sat_catalog y search_sat_units.\n"
            "Estos productos NO fueron clasificados en la primera ronda. "
            "AHORA debes ser más agresivo y persistente:\n"
            "(1) Busca al menos 4 veces por producto con términos distintos: específico, sinónimo, "
            "genérico, y categoría del capítulo HS.\n"
            "(2) Si no encuentras el código perfecto, elige el MÁS CERCANO disponible — "
            "es preferible un código aproximado de la categoría correcta que null.\n"
            "(3) null SOLO si no existe absolutamente ningún código remotamente relacionado en todo el catálogo.\n"
            "(4) Traducciones clave: 'manga/funda para vaso'→busca 'funda','protector','aislante','cubierta'; "
            "'portavaso'→'soporte','bandeja','porta vasos','organizador'; "
            "'tapa domo'→'tapa','cubierta','tapadera','tapa vaso'; "
            "'contenedor aluminio'→'recipiente','envase','contenedor'; "
            "'cubre asiento'→'cubierta sanitaria','protector sanitario','higiene'.\n"
            "(5) Incluye confidence: 'medium' si el código es razonablemente cercano, "
            "'low' si es el más cercano pero puede no ser correcto. Nunca 'high' en esta ronda.\n"
            "(6) Solo responde JSON cuando hayas procesado TODOS los productos de esta lista."
        )

        user_msg2 = (
            f"Estos {len(null_partidas)} productos quedaron sin clasificar. Intenta más fuerte:\n\n"
            f"Productos:\n{_items_text(null_partidas)}\n\n"
            f"Busca cada uno AL MENOS 4 VECES. Elige el código más cercano si no encuentras el exacto.\n"
            f"Responde ÚNICAMENTE con este JSON (sin markdown):\n"
            f'[{{"fraccion":"...","key":"... o null","unit_key":"H87 u otra","description":"... o null","confidence":"medium|low"}}]'
        )

        rescue_json = run_loop(
            [types.Content(role="user", parts=[types.Part(text=user_msg2)])],
            system_pass2,
            len(null_partidas),
        )

        # Merge rescue results into final_json (override nulls with rescued values)
        if rescue_json:
            rescue_map = {item["fraccion"]: item for item in rescue_json}
            for i, item in enumerate(final_json):
                if item["fraccion"] in rescue_map:
                    rescued = rescue_map[item["fraccion"]]
                    # Ensure second-pass items are never marked high confidence
                    if rescued.get("confidence") == "high":
                        rescued["confidence"] = "medium"
                    final_json[i] = rescued

    # ── Save to Productos ─────────────────────────────────────────────────────
    results = []
    for item in final_json:
        frac       = item.get("fraccion", "")
        key        = item.get("key") or None
        desc_sat   = item.get("description") or ""
        unit_key   = (item.get("unit_key") or "H87").strip()
        confidence = item.get("confidence") or "high"
        if confidence not in ("high", "medium", "low"):
            confidence = "high"

        if not key or key.lower() == "null":
            results.append({"fraccion": frac, "key": None, "status": "skipped", "in_catalog": False, "description": None})
            continue

        key = key.strip()
        local_row = session.execute(
            sa_text("SELECT description FROM sat_claves WHERE key = :k"), {"k": key}
        ).fetchone()
        confirmed_desc = local_row[0] if local_row else desc_sat

        # Keys not in catalog are at best medium confidence
        if not local_row and confidence == "high":
            confidence = "medium"

        orig_partida = next((p for p in to_map if p.fraccion == frac), None)
        if not orig_partida:
            continue

        existing = session.exec(select(Producto).where(Producto.fraccion == frac)).first()
        if existing:
            existing.clave_prod_serv = key
            existing.descripcion_sat = confirmed_desc
            existing.unit_key = unit_key
            existing.confidence = confidence
            session.add(existing)
        else:
            session.add(Producto(
                fraccion=frac,
                descripcion=orig_partida.descripcion,
                clave_prod_serv=key,
                descripcion_sat=confirmed_desc,
                unit_key=unit_key,
                confidence=confidence,
            ))
        results.append({
            "fraccion": frac,
            "key": key,
            "status": "saved",
            "in_catalog": local_row is not None,
            "description": confirmed_desc,
            "unit_key": unit_key,
            "confidence": confidence,
        })

    session.commit()
    saved = sum(1 for r in results if r["status"] == "saved")
    return {"mapped": saved, "skipped": len(results) - saved, "results": results}


@app.delete("/pedimentos/{pedimento_id}")
def delete_pedimento(pedimento_id: int, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    pedimento = session.get(Pedimento, pedimento_id)
    if not pedimento:
        raise HTTPException(404, "Pedimento no encontrado")
    for p in pedimento.partidas:
        session.delete(p)
    session.delete(pedimento)
    session.commit()
    return {"ok": True}


# ── Productos ─────────────────────────────────────────────────────────────────

@app.get("/productos")
def list_productos(_: str = Depends(require_auth), session: Session = Depends(get_session)):
    return session.exec(select(Producto)).all()


@app.get("/productos/{fraccion}")
def get_producto(fraccion: str, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    p = session.exec(select(Producto).where(Producto.fraccion == fraccion)).first()
    if not p:
        raise HTTPException(404, "Fracción no encontrada")
    return p


@app.post("/productos", status_code=201)
def create_producto(body: dict, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    existing = session.exec(select(Producto).where(Producto.fraccion == body["fraccion"])).first()
    if existing:
        raise HTTPException(409, "Ya existe un producto con esa fracción")
    p = Producto(
        fraccion=body["fraccion"],
        descripcion=body["descripcion"],
        clave_prod_serv=body["clave_prod_serv"],
        descripcion_sat=body.get("descripcion_sat"),
        unit_key=body.get("unit_key", "H87"),
        confidence=body.get("confidence"),
    )
    session.add(p)
    session.commit()
    session.refresh(p)
    return p


@app.put("/productos/{fraccion}")
def update_producto(fraccion: str, body: dict, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    p = session.exec(select(Producto).where(Producto.fraccion == fraccion)).first()
    if not p:
        raise HTTPException(404, "Fracción no encontrada")
    p.descripcion = body.get("descripcion", p.descripcion)
    p.clave_prod_serv = body.get("clave_prod_serv", p.clave_prod_serv)
    p.descripcion_sat = body.get("descripcion_sat", p.descripcion_sat)
    p.unit_key = body.get("unit_key", p.unit_key)
    if "confidence" in body:
        p.confidence = body["confidence"]
    session.add(p)
    session.commit()
    session.refresh(p)
    return p


@app.delete("/productos/{fraccion}", status_code=204)
def delete_producto(fraccion: str, _: str = Depends(require_auth), session: Session = Depends(get_session)):
    p = session.exec(select(Producto).where(Producto.fraccion == fraccion)).first()
    if not p:
        raise HTTPException(404, "Fracción no encontrada")
    session.delete(p)
    session.commit()


# ── Clientes (FacturAPI proxy) ────────────────────────────────────────────────

@app.get("/clientes")
async def list_clientes(q: str = "", page: int = 1, limit: int = 20,
                        _: str = Depends(require_auth),
                        client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.get("customers", params={"q": q, "page": page, "limit": limit})
    if r.status_code != 200:
        raise HTTPException(r.status_code, r.text)
    return r.json()


@app.post("/clientes", status_code=201)
async def create_cliente(body: dict, _: str = Depends(require_auth), client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.post("customers", json=body)
    if r.status_code not in (200, 201):
        raise HTTPException(r.status_code, r.json().get("message", r.text))
    return r.json()


@app.get("/clientes/{cliente_id}")
async def get_cliente(cliente_id: str, _: str = Depends(require_auth), client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.get(f"customers/{cliente_id}")
    if r.status_code != 200:
        raise HTTPException(r.status_code, r.text)
    return r.json()


@app.put("/clientes/{cliente_id}")
async def update_cliente(cliente_id: str, body: dict,
                         _: str = Depends(require_auth),
                         client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.put(f"customers/{cliente_id}", json=body)
    if r.status_code != 200:
        raise HTTPException(r.status_code, r.json().get("message", r.text))
    return r.json()


# ── Facturas ──────────────────────────────────────────────────────────────────

def _save_factura(inv: dict, pedimento_id: int | None, session: Session) -> Factura:
    """Upsert a FacturAPI invoice object into the local factura table."""
    existing = session.exec(select(Factura).where(Factura.facturapi_id == inv["id"])).first()
    fecha = datetime.fromisoformat(inv.get("date", inv.get("created_at", "")).replace("Z", "+00:00"))
    if existing:
        existing.status = inv.get("status", existing.status)
        existing.cancellation_status = inv.get("cancellation_status") or "none"
        existing.uuid = inv.get("uuid")
        session.add(existing)
        session.commit()
        return existing
    f = Factura(
        facturapi_id=inv["id"],
        uuid=inv.get("uuid"),
        pedimento_id=pedimento_id,
        status=inv.get("status", "valid"),
        cancellation_status=inv.get("cancellation_status") or "none",
        payment_method=inv.get("payment_method", "PUE"),
        total=inv.get("total", 0),
        currency=inv.get("currency", "MXN"),
        customer_name=inv.get("customer", {}).get("legal_name", ""),
        customer_tax_id=inv.get("customer", {}).get("tax_id", ""),
        serie=inv.get("series"),
        folio_number=inv.get("folio_number"),
        fecha=fecha,
    )
    session.add(f)
    session.commit()
    session.refresh(f)
    return f


@app.get("/facturas")
async def list_facturas(q: str = "", page: int = 1, limit: int = 50,
                        payment_method: str = "",
                        _: str = Depends(require_auth),
                        client: httpx.AsyncClient = Depends(get_client),
                        session: Session = Depends(get_session)):
    params: dict = {"page": page, "limit": limit, "type": "I"}
    if q:
        params["q"] = q
    if payment_method:
        params["payment_method"] = payment_method
    async with client:
        r = await client.get("invoices", params=params)
    if r.status_code != 200:
        raise HTTPException(r.status_code, r.text)
    data = r.json()
    return data


@app.post("/facturas", status_code=201)
async def create_factura(body: dict,
                         _: str = Depends(require_auth),
                         client: httpx.AsyncClient = Depends(get_client),
                         session: Session = Depends(get_session)):
    pedimento_id = body.pop("pedimento_id", None)
    async with client:
        r = await client.post("invoices", json=body)
    if r.status_code not in (200, 201):
        detail = r.text
        try:
            detail = r.json().get("message", detail)
        except Exception:
            pass
        raise HTTPException(r.status_code, detail)
    inv = r.json()
    _save_factura(inv, pedimento_id, session)
    return inv


@app.post("/facturas/preview")
async def preview_factura(body: dict, _: str = Depends(require_auth), client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.post("invoices/preview/pdf", json=body)
    if r.status_code != 200:
        detail = r.text
        try:
            detail = r.json().get("message", detail)
        except Exception:
            pass
        raise HTTPException(r.status_code, detail)
    return Response(content=r.content, media_type="application/pdf")


@app.get("/facturas/{factura_id}/pdf")
async def download_pdf(factura_id: str, _: str = Depends(require_auth), client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.get(f"invoices/{factura_id}/pdf")
    if r.status_code != 200:
        raise HTTPException(r.status_code, "No se pudo descargar el PDF")
    return Response(content=r.content, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{factura_id}.pdf"'})


@app.get("/facturas/{factura_id}/xml")
async def download_xml(factura_id: str, _: str = Depends(require_auth), client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.get(f"invoices/{factura_id}/xml")
    if r.status_code != 200:
        raise HTTPException(r.status_code, "No se pudo descargar el XML")
    return Response(content=r.content, media_type="application/xml",
                    headers={"Content-Disposition": f'attachment; filename="{factura_id}.xml"'})


@app.post("/facturas/{factura_id}/email")
async def send_factura_email(factura_id: str, body: dict = {},
                             _: str = Depends(require_auth),
                             client: httpx.AsyncClient = Depends(get_client)):
    async with client:
        r = await client.post(f"invoices/{factura_id}/email", json=body)
    if r.status_code != 200:
        raise HTTPException(r.status_code, "No se pudo enviar el correo")
    return r.json()


@app.delete("/facturas/{factura_id}")
async def cancel_factura(factura_id: str, motive: str = "02",
                         substitution: str = "",
                         _: str = Depends(require_auth),
                         client: httpx.AsyncClient = Depends(get_client),
                         session: Session = Depends(get_session)):
    params: dict = {"motive": motive}
    if substitution:
        params["substitution"] = substitution
    async with client:
        r = await client.delete(f"invoices/{factura_id}", params=params)
    if r.status_code != 200:
        detail = r.text
        try:
            detail = r.json().get("message", detail)
        except Exception:
            pass
        raise HTTPException(r.status_code, detail)
    inv = r.json()
    existing = session.exec(select(Factura).where(Factura.facturapi_id == factura_id)).first()
    if existing:
        existing.status = inv.get("status", "canceled")
        existing.cancellation_status = inv.get("cancellation_status") or "canceled"
        session.add(existing)
        session.commit()
    return inv


# ── Complementos de Pago ──────────────────────────────────────────────────────

@app.get("/complementos")
def list_complementos(_: str = Depends(require_auth), session: Session = Depends(get_session)):
    rows = session.exec(select(ComplementoPago)).all()
    return [
        {
            "id": c.id,
            "facturapi_id": c.facturapi_id,
            "uuid": c.uuid,
            "factura_id": c.factura_id,
            "fecha_pago": c.fecha_pago,
            "monto": c.monto,
            "forma_pago": c.forma_pago,
            "created_at": c.created_at,
        }
        for c in rows
    ]


@app.post("/complementos", status_code=201)
async def create_complemento(body: dict,
                              _: str = Depends(require_auth),
                              client: httpx.AsyncClient = Depends(get_client),
                              session: Session = Depends(get_session)):
    factura_facturapi_id = body["factura_facturapi_id"]
    forma_pago = body["forma_pago"]
    monto = float(body["monto"])
    fecha_pago_str = body["fecha_pago"]  # YYYY-MM-DD

    # Fetch original invoice from FacturAPI to get uuid, customer, and total.
    # The list endpoint accepts exact match on id via the q param; avoids
    # a known httpx base_url path-resolution quirk with the detail endpoint.
    async with client:
        r = await client.get("invoices", params={"q": factura_facturapi_id, "limit": 1})
        if r.status_code != 200:
            raise HTTPException(r.status_code, "Factura no encontrada en FacturAPI")
        results = r.json().get("data", [])
        if not results:
            raise HTTPException(404, "Factura no encontrada en FacturAPI")
        inv = results[0]

        uuid = inv.get("uuid")
        cust = inv.get("customer", {})
        total = float(inv.get("total", monto))

        # Build complemento de pago body
        # IVA 16% assumed — standard for mercancía en México
        iva_base = round(monto / 1.16, 6)
        # Pass customer inline, stripping read-only fields FacturAPI returns
        customer_obj = {k: v for k, v in cust.items()
                        if k not in ("id", "created_at", "updated_at", "livemode")}
        complement_body = {
            "type": "P",
            "customer": customer_obj,
            "complements": [{
                "type": "pago",
                "data": [{
                    "payment_form": forma_pago,
                    "date": f"{fecha_pago_str}T12:00:00",
                    "related_documents": [{
                        "uuid": uuid,
                        "amount": monto,
                        "installment": 1,
                        "last_balance": total,
                        "taxes": [{
                            "base": iva_base,
                            "type": "IVA",
                            "rate": 0.16,
                            "factor": "Tasa",
                            "withholding": False,
                        }],
                        "taxability": "02",
                    }]
                }]
            }]
        }

        cr = await client.post("invoices", json=complement_body)
        if cr.status_code not in (200, 201):
            detail = cr.text
            try:
                detail = cr.json().get("message", detail)
            except Exception:
                pass
            raise HTTPException(cr.status_code, detail)
        comp = cr.json()

    # Ensure factura exists locally (upsert from FacturAPI data)
    local_factura = _save_factura(inv, None, session)

    c = ComplementoPago(
        facturapi_id=comp["id"],
        uuid=comp.get("uuid"),
        factura_id=local_factura.id,
        fecha_pago=date.fromisoformat(fecha_pago_str),
        monto=monto,
        forma_pago=forma_pago,
    )
    session.add(c)
    session.commit()
    session.refresh(c)
    return comp


# ── Catalogs (FacturAPI proxy) ─────────────────────────────────────────────────

@app.get("/catalogs/products")
def catalog_products(q: str = "", _: str = Depends(require_auth), session: Session = Depends(get_session)):
    if not q or len(q) < 2:
        return {"data": []}
    q = q.strip()
    # Key prefix match (e.g. user typed "8471" or "84713")
    if re.match(r'^[0-9A-Z]+$', q):
        rows = session.execute(
            sa_text("SELECT key, description FROM sat_claves WHERE key LIKE :q LIMIT 25"),
            {"q": f"{q}%"},
        ).fetchall()
        return {"data": [{"key": r[0], "description": r[1]} for r in rows]}
    # Full-text search — OR mode for responsive autocomplete
    words = [w for w in re.findall(r"[a-z]{2,}", _norm(q))]
    if not words:
        return {"data": []}
    rows = _fts_search(session, "sat_claves_fts", "sat_claves", words, limit=25, prefer_or=True)
    return {"data": rows}


@app.get("/catalogs/units")
def catalog_units(q: str = "", _: str = Depends(require_auth), session: Session = Depends(get_session)):
    if not q or len(q) < 1:
        return {"data": []}
    q = q.strip()
    # Key prefix match for short alpha-numeric inputs like "H87", "KGM"
    if re.match(r'^[0-9A-Za-z]+$', q) and len(q) <= 5:
        rows = session.execute(
            sa_text("SELECT key, description FROM sat_unidades WHERE key LIKE :q LIMIT 15"),
            {"q": f"{q.upper()}%"},
        ).fetchall()
        if rows:
            return {"data": [{"key": r[0], "description": r[1]} for r in rows]}
    # Full-text search
    words = [w for w in re.findall(r"[a-z]{2,}", _norm(q))]
    if not words:
        return {"data": []}
    rows = _fts_search(session, "sat_unidades_fts", "sat_unidades", words, limit=15, prefer_or=True)
    return {"data": rows}


# ── Export ────────────────────────────────────────────────────────────────────

@app.post("/export")
async def export_xlsx(request: Request, _: str = Depends(require_auth)):
    data = await request.json()
    partidas = data.get("partidas", [])
    tc = data.get("tipo_cambio", 0)
    ped_num = data.get("pedimento_num", "")
    importador = data.get("importador", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Partidas"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="0C1E35")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(bottom=Side(style="thin", color="E2E6ED"))
    amber_fill = PatternFill("solid", fgColor="FEF8EC")
    num_fmt_money = '#,##0.00'
    num_fmt_precise = '#,##0.00000'

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Pedimento: {ped_num}"
    ws["A1"].font = Font(bold=True, size=12)
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Importador: {importador}"
    ws["A2"].font = Font(size=10, color="64748B")

    headers = ["Partida", "Valor de Aduana", "Piezas", "Tipo de Cambio",
               "P.U USD", "Valor Dlls", "P.U MN", "Incrementables"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, p in enumerate(partidas):
        row = 5 + i
        pu_mn = p["val_aduana"] / p["cantidad"] if p["cantidad"] else 0
        pu_usd = pu_mn / tc if tc else 0
        valor_dlls = p["val_aduana"] / tc if tc else 0
        has_inc = p.get("tiene_incrementables", False)

        ws.cell(row=row, column=1, value=p["sec"])
        ws.cell(row=row, column=2, value=p["val_aduana"]).number_format = num_fmt_money
        ws.cell(row=row, column=3, value=p["cantidad"])
        ws.cell(row=row, column=4, value=tc if tc else None).number_format = num_fmt_precise
        ws.cell(row=row, column=5, value=pu_usd if tc else None).number_format = num_fmt_precise
        ws.cell(row=row, column=6, value=valor_dlls if tc else None).number_format = num_fmt_money
        ws.cell(row=row, column=7, value=pu_mn).number_format = num_fmt_precise
        ws.cell(row=row, column=8, value="Sí" if has_inc else "No")

        if has_inc:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = amber_fill
                ws.cell(row=row, column=col).border = thin_border

    col_widths = [10, 16, 10, 16, 16, 14, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pedimento_{ped_num.replace(' ', '_')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Frontend ──────────────────────────────────────────────────────────────────

@app.get("/")
async def index():
    return FileResponse(FRONTEND_DIR / "index.html")


app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")
